from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from html import escape
from io import StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from .. import crud, models

def _to_float(value: Decimal | float | int | None) -> float:
    if value is None:
        return 0.0
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return 0.0
    return round(amount, 2)


def _to_krw_cost(
    value: Decimal | float | int | None,
    currency: str | None,
    usd_krw_rate: float,
) -> float:
    return round(float(crud._cost_to_krw(value, currency, usd_krw_rate)), 2)


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _to_date_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return ""

    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        return text


def _normalize_department(value: str | None) -> str:
    text = str(value or "").strip()
    return text or "미할당"


def _resolve_team_bucket_key_for_report(org_unit_id: int | None, department: str | None) -> str:
    if org_unit_id is not None:
        try:
            normalized_id = int(org_unit_id)
        except (TypeError, ValueError):
            normalized_id = 0
        if normalized_id > 0:
            return f"org:{normalized_id}"

    return f"dept:{_normalize_department(department)}"


def _resolve_team_name_for_report(org_unit_name: str | None, department: str | None) -> str:
    # Phase 2 preparation: report display is fallback-compatible while grouping key can move to org_unit_id-first.
    return _normalize_department(org_unit_name or department)


def _resolve_team_identity_for_report(
    org_unit_id: int | None,
    org_unit_name: str | None,
    department: str | None,
) -> tuple[str, str]:
    return (
        _resolve_team_bucket_key_for_report(org_unit_id, department),
        _resolve_team_name_for_report(org_unit_name, department),
    )


def _coerce_detail_item(raw: Any) -> dict[str, Any]:
    if isinstance(raw, dict):
        return raw
    model_dump = getattr(raw, "model_dump", None)
    if callable(model_dump):
        data = model_dump()
        if isinstance(data, dict):
            return data
    return {}


def _build_directory_user_maps(db: Session) -> tuple[dict[str, str], dict[str, str]]:
    department_map: dict[str, str] = {}
    display_name_map: dict[str, str] = {}

    org_name_by_id: dict[int, str] = {}
    for org_id, org_name in db.query(models.OrganizationUnit.id, models.OrganizationUnit.name).all():
        key = int(org_id)
        name = str(org_name or "").strip()
        if not name:
            continue
        org_name_by_id[key] = name

    rows = db.query(
        models.DirectoryUser.username,
        models.DirectoryUser.display_name,
        models.DirectoryUser.department,
        models.DirectoryUser.org_unit_id,
    ).all()

    for username, display_name, department, org_unit_id in rows:
        key = str(username or "").strip()
        if not key:
            continue

        display_text = str(display_name or "").strip() or key
        display_name_map[key] = display_text

        normalized_org_id = int(org_unit_id) if org_unit_id else None
        org_name = org_name_by_id.get(normalized_org_id) if normalized_org_id else None
        _team_key, team_name = _resolve_team_identity_for_report(normalized_org_id, org_name, department)
        department_map[key] = team_name

    return department_map, display_name_map


def _build_directory_user_team_identity_map(db: Session) -> dict[str, dict[str, str]]:
    team_identity_map: dict[str, dict[str, str]] = {}

    org_name_by_id: dict[int, str] = {}
    for org_id, org_name in db.query(models.OrganizationUnit.id, models.OrganizationUnit.name).all():
        key = int(org_id)
        name = str(org_name or "").strip()
        if not name:
            continue
        org_name_by_id[key] = name

    rows = db.query(
        models.DirectoryUser.username,
        models.DirectoryUser.department,
        models.DirectoryUser.org_unit_id,
    ).all()

    for username, department, org_unit_id in rows:
        key = str(username or "").strip()
        if not key:
            continue

        normalized_org_id = int(org_unit_id) if org_unit_id else None
        org_name = org_name_by_id.get(normalized_org_id) if normalized_org_id else None
        bucket_key, team_name = _resolve_team_identity_for_report(normalized_org_id, org_name, department)
        team_identity_map[key] = {
            "bucket_key": bucket_key,
            "team_name": team_name,
        }

    return team_identity_map


def _extract_license_assignees(
    license_row: models.SoftwareLicense,
    department_map: dict[str, str],
    display_name_map: dict[str, str],
) -> list[dict[str, str]]:
    detail_map: dict[str, dict[str, Any]] = {}
    details = license_row.assignee_details if isinstance(license_row.assignee_details, list) else []

    for raw in details:
        item = _coerce_detail_item(raw)
        username = str(item.get("username") or "").strip()
        if not username or username in detail_map:
            continue
        detail_map[username] = item

    usernames: list[str] = []
    seen: set[str] = set()

    assignees = license_row.assignees if isinstance(license_row.assignees, list) else []
    for raw in assignees:
        username = str(raw or "").strip()
        if not username or username in seen:
            continue
        seen.add(username)
        usernames.append(username)

    for username in detail_map:
        if username in seen:
            continue
        seen.add(username)
        usernames.append(username)

    rows: list[dict[str, str]] = []
    default_start_date = _to_date_text(license_row.start_date)
    default_end_date = _to_date_text(license_row.end_date)

    for username in usernames:
        detail = detail_map.get(username, {})
        display_name = display_name_map.get(username, username)
        department = department_map.get(username, "미할당")
        start_date = _to_date_text(detail.get("start_date")) or default_start_date
        end_date = _to_date_text(detail.get("end_date")) or default_end_date

        rows.append(
            {
                "username": username,
                "display_name": display_name,
                "department": department,
                "start_date": start_date,
                "end_date": end_date,
            }
        )

    return rows


def _monthly_unit_cost(subscription_type: str, unit_cost_krw: float) -> float:
    if str(subscription_type or "").strip() == "연 구독":
        return round(unit_cost_krw / 12.0, 2)
    return round(unit_cost_krw, 2)


def _normalize_scope_filter(scope_filter: str | None) -> str:
    key = str(scope_filter or "").strip().lower()
    if key in {"required", "필수", "mandatory", "critical"}:
        return "required"
    if key in {"general", "일반"}:
        return "general"
    return "all"


def _match_scope_filter(scope_filter: str, normalized_scope: str) -> bool:
    if scope_filter == "required":
        return normalized_scope == "필수"
    if scope_filter == "general":
        return normalized_scope == "일반"
    return True


def build_dashboard_software_cost_summary(db: Session, scope_filter: str = "all") -> dict[str, Any]:
    normalized_filter = _normalize_scope_filter(scope_filter)

    department_map, display_name_map = _build_directory_user_maps(db)
    team_identity_map = _build_directory_user_team_identity_map(db)
    exchange_rate_setting = crud.get_exchange_rate_setting(db)
    usd_krw_rate = float(exchange_rate_setting.get("usd_krw") or crud.DEFAULT_USD_KRW_RATE)

    licenses = (
        db.query(models.SoftwareLicense)
        .order_by(models.SoftwareLicense.product_name.asc(), models.SoftwareLicense.id.asc())
        .all()
    )

    team_buckets: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "team_name": "\ubbf8\ud560\ub2f9",
            "users": set(),
            "assigned_license_count": 0,
            "monthly_cost": 0.0,
            "license_types": set(),
        }
    )

    overall_users: set[str] = set()
    overall_assigned_license_count = 0
    overall_monthly_cost = 0.0
    overall_license_types: set[int] = set()

    for license_row in licenses:
        normalized_scope = crud.normalize_license_scope(getattr(license_row, "license_scope", None))
        if not _match_scope_filter(normalized_filter, normalized_scope):
            continue

        unit_cost_krw = _to_krw_cost(
            license_row.purchase_cost,
            license_row.purchase_currency,
            usd_krw_rate,
        )
        if unit_cost_krw <= 0:
            continue

        monthly_unit_cost = _monthly_unit_cost(str(license_row.subscription_type or ""), unit_cost_krw)
        assignee_rows = _extract_license_assignees(license_row, department_map, display_name_map)
        if not assignee_rows:
            continue

        assignment_count_by_user: dict[str, int] = defaultdict(int)
        raw_assignees = [
            str(value or "").strip()
            for value in (license_row.assignees if isinstance(license_row.assignees, list) else [])
            if str(value or "").strip()
        ]
        for username in raw_assignees:
            assignment_count_by_user[username] += 1

        if not assignment_count_by_user:
            for row in assignee_rows:
                username = str(row.get("username") or "").strip()
                if username:
                    assignment_count_by_user[username] += 1

        if not assignment_count_by_user:
            continue

        license_id = _to_int(getattr(license_row, "id", 0), default=0)

        for username, seat_count in assignment_count_by_user.items():
            count = max(0, _to_int(seat_count, default=0))
            if count <= 0:
                continue

            identity = team_identity_map.get(username)
            if identity:
                team_name = _normalize_department(identity.get("team_name"))
                bucket_key = str(identity.get("bucket_key") or f"dept:{team_name}")
            else:
                team_name = _normalize_department(department_map.get(username))
                bucket_key = f"dept:{team_name}"

            bucket = team_buckets[bucket_key]
            bucket["team_name"] = team_name
            bucket["users"].add(username)
            bucket["assigned_license_count"] += count
            bucket["monthly_cost"] += float(monthly_unit_cost) * float(count)
            if license_id > 0:
                bucket["license_types"].add(license_id)
                overall_license_types.add(license_id)

            overall_users.add(username)
            overall_assigned_license_count += count
            overall_monthly_cost += float(monthly_unit_cost) * float(count)

    team_summary = [
        {
            "team_name": str(bucket.get("team_name") or "\ubbf8\ud560\ub2f9"),
            "user_count": len(bucket["users"]),
            "assigned_license_count": _to_int(bucket["assigned_license_count"]),
            "monthly_cost": round(_to_float(bucket["monthly_cost"]), 2),
            "yearly_cost": round(_to_float(bucket["monthly_cost"]) * 12.0, 2),
            "license_type_count": len(bucket["license_types"]),
        }
        for bucket in team_buckets.values()
    ]
    team_summary.sort(key=lambda row: (-float(row["monthly_cost"]), row["team_name"]))

    overall_summary = {
        "team_count": len(team_summary),
        "user_count": len(overall_users),
        "assigned_license_count": overall_assigned_license_count,
        "monthly_cost": round(overall_monthly_cost, 2),
        "yearly_cost": round(overall_monthly_cost * 12.0, 2),
        "license_type_count": len(overall_license_types),
    }

    return {
        "scope_filter": normalized_filter,
        "overall_summary": overall_summary,
        "team_summary": team_summary,
    }

def _normalize_snapshot_month(value: date | None) -> date:
    target = value or date.today()
    return date(target.year, target.month, 1)


def create_software_cost_snapshot(
    db: Session,
    *,
    snapshot_month: date | None = None,
    scope_filter: str = "all",
    overwrite: bool = False,
) -> dict[str, Any]:
    month = _normalize_snapshot_month(snapshot_month)
    normalized_filter = _normalize_scope_filter(scope_filter)

    existing_query = db.query(models.SoftwareCostSnapshot).filter(
        models.SoftwareCostSnapshot.snapshot_month == month,
        models.SoftwareCostSnapshot.scope == normalized_filter,
    )
    has_existing = existing_query.first() is not None
    overwritten = False

    if has_existing and not overwrite:
        raise ValueError("snapshot_exists")

    if has_existing and overwrite:
        existing_query.delete(synchronize_session=False)
        overwritten = True

    summary = build_dashboard_software_cost_summary(db, scope_filter=normalized_filter)
    team_summary = summary.get("team_summary") if isinstance(summary, dict) else []
    team_rows = team_summary if isinstance(team_summary, list) else []

    # TODO(phase2): snapshot? team_name ? org identity(bucket key) ?? ?? ?? ??
    rows_to_create: list[models.SoftwareCostSnapshot] = []
    for row in team_rows:
        if not isinstance(row, dict):
            continue
        team_name = _normalize_department(str(row.get("team_name") or ""))
        user_count = _to_int(row.get("user_count"), default=0)
        license_count = _to_int(row.get("assigned_license_count"), default=0)
        monthly_cost = _to_float(row.get("monthly_cost"))
        annual_cost = _to_float(row.get("yearly_cost"))

        rows_to_create.append(
            models.SoftwareCostSnapshot(
                snapshot_month=month,
                team_name=team_name,
                scope=normalized_filter,
                user_count=max(0, user_count),
                license_count=max(0, license_count),
                monthly_cost=monthly_cost,
                annual_cost=annual_cost,
            )
        )

    if rows_to_create:
        db.add_all(rows_to_create)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise ValueError("snapshot_exists")

    rows = (
        db.query(models.SoftwareCostSnapshot)
        .filter(
            models.SoftwareCostSnapshot.snapshot_month == month,
            models.SoftwareCostSnapshot.scope == normalized_filter,
        )
        .order_by(models.SoftwareCostSnapshot.monthly_cost.desc(), models.SoftwareCostSnapshot.team_name.asc())
        .all()
    )

    return {
        "snapshot_month": month,
        "scope_filter": normalized_filter,
        "overwritten": overwritten,
        "created_count": len(rows),
        "rows": rows,
    }


def list_software_cost_snapshots(
    db: Session,
    *,
    scope_filter: str = "all",
    snapshot_month_from: date | None = None,
    snapshot_month_to: date | None = None,
    limit: int = 500,
) -> dict[str, Any]:
    normalized_filter = _normalize_scope_filter(scope_filter)
    safe_limit = max(1, min(limit, 5000))

    query = db.query(models.SoftwareCostSnapshot)
    if normalized_filter != "all":
        query = query.filter(models.SoftwareCostSnapshot.scope == normalized_filter)

    month_from = _normalize_snapshot_month(snapshot_month_from) if snapshot_month_from else None
    month_to = _normalize_snapshot_month(snapshot_month_to) if snapshot_month_to else None

    if month_from:
        query = query.filter(models.SoftwareCostSnapshot.snapshot_month >= month_from)
    if month_to:
        query = query.filter(models.SoftwareCostSnapshot.snapshot_month <= month_to)

    total = query.count()
    rows = (
        query.order_by(
            models.SoftwareCostSnapshot.snapshot_month.desc(),
            models.SoftwareCostSnapshot.monthly_cost.desc(),
            models.SoftwareCostSnapshot.team_name.asc(),
            models.SoftwareCostSnapshot.id.asc(),
        )
        .limit(safe_limit)
        .all()
    )

    return {
        "scope_filter": normalized_filter,
        "snapshot_month_from": month_from,
        "snapshot_month_to": month_to,
        "total": total,
        "rows": rows,
    }


def build_general_license_report_data(db: Session) -> dict[str, Any]:
    department_map, display_name_map = _build_directory_user_maps(db)
    team_identity_map = _build_directory_user_team_identity_map(db)
    exchange_rate_setting = crud.get_exchange_rate_setting(db)
    usd_krw_rate = float(exchange_rate_setting.get("usd_krw") or crud.DEFAULT_USD_KRW_RATE)

    general_licenses = [
        row
        for row in db.query(models.SoftwareLicense).order_by(models.SoftwareLicense.product_name.asc(), models.SoftwareLicense.id.asc()).all()
        if crud.normalize_license_scope(getattr(row, "license_scope", None)) == "\uc77c\ubc18"
    ]

    team_buckets: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"team_name": "\ubbf8\ud560\ub2f9", "users": set(), "assigned_quantity": 0, "monthly_cost": 0.0}
    )
    license_summary: list[dict[str, Any]] = []
    user_detail: list[dict[str, Any]] = []
    all_assigned_users: set[str] = set()

    total_cost = 0.0
    total_license_quantity = 0

    for license_row in general_licenses:
        product_name = str(license_row.product_name or "(\uc774\ub984\uc5c6\uc74c)").strip() or "(\uc774\ub984\uc5c6\uc74c)"
        total_quantity = max(0, _to_int(license_row.total_quantity, default=0))

        unit_cost_krw = _to_krw_cost(license_row.purchase_cost, license_row.purchase_currency, usd_krw_rate)
        total_period_cost = round(unit_cost_krw * total_quantity, 2)
        total_cost += total_period_cost
        total_license_quantity += total_quantity

        subscription_type = str(license_row.subscription_type or "").strip()
        monthly_unit_cost = _monthly_unit_cost(subscription_type, unit_cost_krw)

        assignee_rows = _extract_license_assignees(license_row, department_map, display_name_map)

        raw_assignees = [
            str(value or "").strip()
            for value in (license_row.assignees if isinstance(license_row.assignees, list) else [])
            if str(value or "").strip()
        ]

        assignment_count_by_user: dict[str, int] = defaultdict(int)
        for username in raw_assignees:
            assignment_count_by_user[username] += 1

        if not assignment_count_by_user and assignee_rows:
            for row in assignee_rows:
                username = str(row.get("username") or "").strip()
                if username:
                    assignment_count_by_user[username] += 1

        unique_assigned_users: set[str] = set(assignment_count_by_user.keys())
        for row in assignee_rows:
            username = str(row.get("username") or "").strip()
            if username:
                unique_assigned_users.add(username)
        all_assigned_users.update(unique_assigned_users)

        team_users_map: dict[str, set[str]] = defaultdict(set)
        team_name_by_bucket: dict[str, str] = {}
        for username in unique_assigned_users:
            identity = team_identity_map.get(username)
            if identity:
                bucket_key = str(identity.get("bucket_key") or "")
                team_name = _normalize_department(identity.get("team_name"))
            else:
                team_name = _normalize_department(department_map.get(username))
                bucket_key = f"dept:{team_name}"

            if not bucket_key:
                bucket_key = f"dept:{team_name}"
            team_users_map[bucket_key].add(username)
            team_name_by_bucket[bucket_key] = team_name

        team_assignment_counts: dict[str, int] = defaultdict(int)
        for username, count in assignment_count_by_user.items():
            identity = team_identity_map.get(username)
            if identity:
                bucket_key = str(identity.get("bucket_key") or "")
                team_name = _normalize_department(identity.get("team_name"))
            else:
                team_name = _normalize_department(department_map.get(username))
                bucket_key = f"dept:{team_name}"

            if not bucket_key:
                bucket_key = f"dept:{team_name}"
            team_name_by_bucket[bucket_key] = team_name
            team_assignment_counts[bucket_key] += int(count)

        if team_assignment_counts:
            for bucket_key, seat_count in team_assignment_counts.items():
                bucket = team_buckets[bucket_key]
                bucket["team_name"] = team_name_by_bucket.get(bucket_key, "\ubbf8\ud560\ub2f9")
                bucket["users"].update(team_users_map.get(bucket_key, set()))
                bucket["assigned_quantity"] += int(seat_count)
                bucket["monthly_cost"] += float(monthly_unit_cost) * float(seat_count)

        teams_for_license = sorted({team_name_by_bucket.get(k, "???") for k in team_users_map.keys()}) if team_users_map else ["\ubbf8\ud560\ub2f9"]

        license_summary.append(
            {
                "license_name": product_name,
                "team": ", ".join(teams_for_license),
                "quantity": total_quantity,
                "user_count": len(unique_assigned_users),
                "cost": total_period_cost,
                "end_date": _to_date_text(license_row.end_date),
            }
        )

        for row in assignee_rows:
            username = str(row.get("username") or "").strip()
            owned_quantity = max(1, int(assignment_count_by_user.get(username, 1)))
            identity = team_identity_map.get(username)
            if identity:
                team_name = _normalize_department(identity.get("team_name"))
            else:
                team_name = _normalize_department(row.get("department"))

            user_detail.append(
                {
                    "user": row["display_name"],
                    "team": team_name,
                    "license_name": product_name,
                    "unit_cost": unit_cost_krw,
                    "owned_quantity": owned_quantity,
                    "start_date": row.get("start_date") or "",
                    "end_date": row.get("end_date") or "",
                    "review_status": "",
                    "note": "",
                }
            )

    team_summary = [
        {
            "team_name": str(bucket.get("team_name") or "\ubbf8\ud560\ub2f9"),
            "user_count": len(bucket["users"]),
            "license_count": int(bucket["assigned_quantity"]),
            "monthly_cost": round(float(bucket["monthly_cost"]), 2),
        }
        for bucket in team_buckets.values()
    ]
    team_summary.sort(key=lambda row: row["team_name"])

    license_summary.sort(key=lambda row: (row["license_name"], row["team"]))
    user_detail.sort(key=lambda row: (row["team"], row["user"], row["license_name"]))

    summary = {
        "\uae30\uc900\uc77c": date.today().isoformat(),
        "\ucd1d \ube44\uc6a9": round(total_cost, 2),
        "\ucd1d \uc0ac\uc6a9\uc790 \uc218": len(all_assigned_users),
        "\ucd1d \ub77c\uc774\uc120\uc2a4 \uc218": total_license_quantity,
    }

    return {
        "summary": summary,
        "team_summary": team_summary,
        "license_summary": license_summary,
        "user_detail": user_detail,
    }

def _apply_header_style(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _apply_cost_number_formats(ws) -> None:
    cost_columns: list[int] = []
    for col_idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value or "")
        if "비용" in header:
            cost_columns.append(col_idx)

    for col_idx in cost_columns:
        for row_idx in range(2, ws.max_row + 1):
            target = ws.cell(row=row_idx, column=col_idx)
            if isinstance(target.value, (int, float)):
                target.number_format = "#,##0.00"


def _apply_summary_value_cost_formats(ws) -> None:
    for row_idx in range(2, ws.max_row + 1):
        label = str(ws.cell(row=row_idx, column=1).value or "")
        value_cell = ws.cell(row=row_idx, column=2)
        if "비용" in label and isinstance(value_cell.value, (int, float)):
            value_cell.number_format = "#,##0.00"


def _adjust_column_widths(ws, min_width: int = 10, max_width: int = 48) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
        width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(column_index)].width = width


def create_general_license_report_workbook(report_data: dict[str, Any]) -> Workbook:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["항목", "값"])
    summary = report_data.get("summary") or {}
    ws_summary.append(["기준일", summary.get("기준일", "")])
    ws_summary.append(["총 비용(원화 환산, 1인 단가 * 총 보유 수량 기준)", summary.get("총 비용", 0)])
    ws_summary.append(["총 사용자 수", summary.get("총 사용자 수", 0)])
    ws_summary.append(["총 라이선스 수", summary.get("총 라이선스 수", 0)])
    ws_summary.freeze_panes = "A2"
    _apply_header_style(ws_summary)
    _apply_cost_number_formats(ws_summary)
    _apply_summary_value_cost_formats(ws_summary)
    _adjust_column_widths(ws_summary)

    ws_team = wb.create_sheet("Team Summary")
    ws_team.append(["팀명", "사용자 수", "라이선스 수", "월 비용"])
    for row in report_data.get("team_summary") or []:
        ws_team.append(
            [
                row.get("team_name", "미할당"),
                row.get("user_count", 0),
                row.get("license_count", 0),
                row.get("monthly_cost", 0),
            ]
        )
    ws_team.freeze_panes = "A2"
    _apply_header_style(ws_team)
    _apply_cost_number_formats(ws_team)
    _adjust_column_widths(ws_team)

    ws_license = wb.create_sheet("License Summary")
    ws_license.append(["라이선스명", "팀", "수량", "사용자 수", "비용", "만료일"])
    for row in report_data.get("license_summary") or []:
        ws_license.append(
            [
                row.get("license_name", ""),
                row.get("team", ""),
                row.get("quantity", 0),
                row.get("user_count", 0),
                row.get("cost", 0),
                row.get("end_date", ""),
            ]
        )
    ws_license.freeze_panes = "A2"
    _apply_header_style(ws_license)
    _apply_cost_number_formats(ws_license)
    _adjust_column_widths(ws_license)

    ws_user = wb.create_sheet("User Detail")
    ws_user.append(["사용자", "팀", "라이선스", "단위비용", "보유수량", "시작일", "종료일", "검토상태", "비고"])
    for row in report_data.get("user_detail") or []:
        ws_user.append(
            [
                row.get("user", ""),
                row.get("team", "미할당"),
                row.get("license_name", ""),
                row.get("unit_cost", 0),
                row.get("owned_quantity", 0),
                row.get("start_date", ""),
                row.get("end_date", ""),
                row.get("review_status", ""),
                row.get("note", ""),
            ]
        )
    ws_user.freeze_panes = "A2"
    _apply_header_style(ws_user)
    _apply_cost_number_formats(ws_user)
    _adjust_column_widths(ws_user)

    return wb



def _format_report_number(value: Any) -> str:
    amount = _to_float(value)
    rounded = round(amount)
    if abs(amount - float(rounded)) < 0.005:
        return f"{rounded:,}"
    return f"{amount:,.2f}"


def create_general_license_report_csv(report_data: dict[str, Any]) -> str:
    output = StringIO(newline="")
    writer = csv.writer(output)

    summary = report_data.get("summary") or {}
    writer.writerow(["Summary"])
    writer.writerow(["항목", "값"])
    writer.writerow(["기준일", summary.get("기준일", "")])
    writer.writerow(["총 비용(원화 환산, 1인 단가 * 총 보유 수량 기준)", _format_report_number(summary.get("총 비용", 0))])
    writer.writerow(["총 사용자 수", _format_report_number(summary.get("총 사용자 수", 0))])
    writer.writerow(["총 라이선스 수", _format_report_number(summary.get("총 라이선스 수", 0))])
    writer.writerow([])

    writer.writerow(["Team Summary"])
    writer.writerow(["팀명", "사용자 수", "라이선스 수", "월 비용"])
    for row in report_data.get("team_summary") or []:
        writer.writerow(
            [
                row.get("team_name", "미할당"),
                _format_report_number(row.get("user_count", 0)),
                _format_report_number(row.get("license_count", 0)),
                _format_report_number(row.get("monthly_cost", 0)),
            ]
        )
    writer.writerow([])

    writer.writerow(["License Summary"])
    writer.writerow(["라이선스명", "팀", "수량", "사용자 수", "비용", "만료일"])
    for row in report_data.get("license_summary") or []:
        writer.writerow(
            [
                row.get("license_name", ""),
                row.get("team", ""),
                _format_report_number(row.get("quantity", 0)),
                _format_report_number(row.get("user_count", 0)),
                _format_report_number(row.get("cost", 0)),
                row.get("end_date", ""),
            ]
        )
    writer.writerow([])

    writer.writerow(["User Detail"])
    writer.writerow(["사용자", "팀", "라이선스", "단위비용", "보유수량", "시작일", "종료일", "검토상태", "비고"])
    for row in report_data.get("user_detail") or []:
        writer.writerow(
            [
                row.get("user", ""),
                row.get("team", "미할당"),
                row.get("license_name", ""),
                _format_report_number(row.get("unit_cost", 0)),
                _format_report_number(row.get("owned_quantity", 0)),
                row.get("start_date", ""),
                row.get("end_date", ""),
                row.get("review_status", ""),
                row.get("note", ""),
            ]
        )

    return output.getvalue()


def create_general_license_report_html(report_data: dict[str, Any]) -> str:
    summary = report_data.get("summary") or {}
    team_rows = report_data.get("team_summary") or []
    license_rows = report_data.get("license_summary") or []
    user_rows = report_data.get("user_detail") or []

    def e(value: Any) -> str:
        return escape(str(value or ""))

    summary_cards = [
        ("기준일", summary.get("기준일", "")),
        ("총 비용", f"{_format_report_number(summary.get('총 비용', 0))} 원"),
        ("총 사용자 수", f"{_format_report_number(summary.get('총 사용자 수', 0))} 명"),
        ("총 라이선스 수", f"{_format_report_number(summary.get('총 라이선스 수', 0))} 개"),
    ]

    team_tbody = "".join(
        f"<tr><td>{e(row.get('team_name', '미할당'))}</td><td>{e(_format_report_number(row.get('user_count', 0)))}</td><td>{e(_format_report_number(row.get('license_count', 0)))}</td><td>{e(_format_report_number(row.get('monthly_cost', 0)))} 원</td></tr>"
        for row in team_rows
    ) or '<tr><td colspan="4" class="empty">데이터가 없습니다.</td></tr>'

    license_tbody = "".join(
        f"<tr><td>{e(row.get('license_name', ''))}</td><td>{e(row.get('team', ''))}</td><td>{e(_format_report_number(row.get('quantity', 0)))}</td><td>{e(_format_report_number(row.get('user_count', 0)))}</td><td>{e(_format_report_number(row.get('cost', 0)))} 원</td><td>{e(row.get('end_date', ''))}</td></tr>"
        for row in license_rows
    ) or '<tr><td colspan="6" class="empty">데이터가 없습니다.</td></tr>'

    user_tbody = "".join(
        f"<tr><td>{e(row.get('user', ''))}</td><td>{e(row.get('team', '미할당'))}</td><td>{e(row.get('license_name', ''))}</td><td>{e(_format_report_number(row.get('unit_cost', 0)))} 원</td><td>{e(_format_report_number(row.get('owned_quantity', 0)))} 개</td><td>{e(row.get('start_date', ''))}</td><td>{e(row.get('end_date', ''))}</td><td>{e(row.get('review_status', ''))}</td><td>{e(row.get('note', ''))}</td></tr>"
        for row in user_rows
    ) or '<tr><td colspan="9" class="empty">데이터가 없습니다.</td></tr>'

    team_user_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in user_rows:
        team_name = str(row.get("team") or "미할당").strip() or "미할당"
        team_user_rows[team_name].append(row)

    team_tab_buttons: list[str] = []
    team_tab_panels: list[str] = []

    for idx, team_name in enumerate(sorted(team_user_rows.keys())):
        panel_id = f"team-{idx}"
        rows = team_user_rows[team_name]
        team_sheet_tbody = "".join(
            f"<tr><td>{e(row.get('user', ''))}</td><td>{e(row.get('license_name', ''))}</td><td>{e(_format_report_number(row.get('owned_quantity', 0)))} 개</td><td>{e(_format_report_number(row.get('unit_cost', 0)))} 원</td><td>{e(row.get('start_date', ''))}</td><td>{e(row.get('end_date', ''))}</td></tr>"
            for row in rows
        ) or '<tr><td colspan="6" class="empty">데이터가 없습니다.</td></tr>'

        team_tab_buttons.append(
            f'<button type="button" class="subtab-btn{' active' if idx == 0 else ''}" data-team-tab-target="{panel_id}">{e(team_name)}</button>'
        )
        team_tab_panels.append(
            f"""
      <section class=\"team-panel{' active' if idx == 0 else ''}\" data-team-panel=\"{panel_id}\">
        <h3 class=\"team-title\">{e(team_name)}</h3>
        <div class=\"table-wrap\">
          <table>
            <thead><tr><th>사용자</th><th>라이선스</th><th>보유수량</th><th>단위비용</th><th>시작일</th><th>종료일</th></tr></thead>
            <tbody>{team_sheet_tbody}</tbody>
          </table>
        </div>
      </section>
"""
        )

    team_tab_buttons_html = "".join(team_tab_buttons)
    team_tab_panels_html = "".join(team_tab_panels)

    cards_html = "".join(
        f"<article class=\"summary-card\"><h3>{e(label)}</h3><p>{e(value)}</p></article>"
        for label, value in summary_cards
    )

    return f"""<!doctype html>
<html lang=\"ko\">
<head>
  <meta charset=\"UTF-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\" />
  <title>일반 라이선스/구독 현황 보고서</title>
  <style>
    :root {{
      --bg: #f6f8fb;
      --card: #ffffff;
      --line: #d8dee9;
      --title: #1f2d3d;
      --text: #2f3b4c;
      --muted: #5f6f82;
      --accent: #1f628f;
      --accent-soft: #e8f1fb;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: var(--bg); color: var(--text); font-family: "Pretendard", "Noto Sans KR", "Segoe UI", sans-serif; }}
    .container {{ max-width: 1320px; margin: 0 auto; padding: 24px; }}
    .header {{ display: flex; align-items: flex-end; justify-content: space-between; gap: 12px; margin-bottom: 16px; }}
    .title {{ margin: 0; font-size: 28px; color: var(--title); }}
    .subtitle {{ margin: 6px 0 0; color: var(--muted); }}
    .tab-nav {{ display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 14px; }}
    .tab-btn {{ border: 1px solid #aac3da; background: #fff; color: #1f4f78; border-radius: 10px; padding: 8px 14px; font-weight: 700; cursor: pointer; }}
    .tab-btn.active {{ background: var(--accent); color: #fff; border-color: var(--accent); }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin: 0 0 20px; }}
    .summary-card {{ background: var(--card); border: 1px solid var(--line); border-radius: 12px; padding: 14px 16px; }}
    .summary-card h3 {{ margin: 0; font-size: 13px; color: var(--muted); font-weight: 700; }}
    .summary-card p {{ margin: 8px 0 0; font-size: 24px; line-height: 1.2; color: var(--title); font-weight: 800; }}
    .section {{ background: var(--card); border: 1px solid var(--line); border-radius: 12px; padding: 14px; margin-bottom: 14px; }}
    .section h2 {{ margin: 0 0 10px; color: var(--title); font-size: 18px; }}
    .table-wrap {{ overflow-x: auto; }}
    table {{ width: 100%; border-collapse: collapse; min-width: 860px; }}
    th, td {{ border: 1px solid var(--line); padding: 8px 10px; font-size: 13px; text-align: left; vertical-align: middle; }}
    thead th {{ background: #eef3f9; color: var(--title); font-weight: 700; position: sticky; top: 0; z-index: 1; }}
    tbody tr:nth-child(even) {{ background: #fbfcff; }}
    .empty {{ text-align: center; color: var(--muted); }}
    .footer-note {{ color: var(--muted); font-size: 12px; margin-top: 8px; }}
    .mono {{ font-family: "Consolas", "Menlo", monospace; }}
    .subtab-nav {{ display: flex; gap: 6px; flex-wrap: wrap; margin-bottom: 10px; }}
    .subtab-btn {{ border: 1px solid #b6c8da; background: #fff; color: #2a4f72; border-radius: 8px; padding: 6px 10px; font-weight: 700; cursor: pointer; font-size: 12px; }}
    .subtab-btn.active {{ background: var(--accent-soft); border-color: #86abd0; color: #1f4f78; }}
    .team-panel {{ display: none; }}
    .team-panel.active {{ display: block; }}
    .team-title {{ margin: 0 0 10px; font-size: 16px; color: var(--title); }}
  </style>
</head>
<body>
  <main class=\"container\">
    <header class=\"header\">
      <div>
        <h1 class=\"title\">일반 라이선스/구독 현황 보고서</h1>
        <p class=\"subtitle\">기준일 <span class=\"mono\">{e(summary.get('기준일', ''))}</span> / 내부 검토용 HTML 리포트</p>
      </div>
    </header>

    <nav class=\"tab-nav\" aria-label=\"보고서 탭\">
      <button type=\"button\" class=\"tab-btn active\" data-tab-target=\"summary\">전체 Summary</button>
      <button type=\"button\" class=\"tab-btn\" data-tab-target=\"user-detail\">User Detail</button>
      <button type=\"button\" class=\"tab-btn\" data-tab-target=\"team-detail\">Team Detail</button>
    </nav>

    <section class=\"tab-panel active\" data-tab-panel=\"summary\">
      <section class=\"summary-grid\">
        {cards_html}
      </section>

      <section class=\"section\">
        <h2>Team Summary</h2>
        <div class=\"table-wrap\">
          <table>
            <thead><tr><th>팀명</th><th>사용자 수</th><th>라이선스 수</th><th>월 비용</th></tr></thead>
            <tbody>{team_tbody}</tbody>
          </table>
        </div>
      </section>

      <section class=\"section\">
        <h2>License Summary</h2>
        <div class=\"table-wrap\">
          <table>
            <thead><tr><th>라이선스명</th><th>팀</th><th>수량</th><th>사용자 수</th><th>비용</th><th>만료일</th></tr></thead>
            <tbody>{license_tbody}</tbody>
          </table>
        </div>
      </section>
    </section>

    <section class=\"tab-panel\" data-tab-panel=\"user-detail\">
      <section class=\"section\">
        <h2>User Detail</h2>
        <div class=\"table-wrap\">
          <table>
            <thead><tr><th>사용자</th><th>팀</th><th>라이선스</th><th>단위비용</th><th>보유수량</th><th>시작일</th><th>종료일</th><th>검토상태</th><th>비고</th></tr></thead>
            <tbody>{user_tbody}</tbody>
          </table>
        </div>
        <p class=\"footer-note\">비용은 설정 환율이 반영된 원화 기준입니다.</p>
      </section>
    </section>

    <section class=\"tab-panel\" data-tab-panel=\"team-detail\">
      <section class=\"section\">
        <h2>Team Detail</h2>
        {f'<div class="subtab-nav">{team_tab_buttons_html}</div><div>{team_tab_panels_html}</div>' if team_tab_buttons_html else '<div class="empty">팀별 상세 데이터가 없습니다.</div>'}
      </section>
    </section>
  </main>

  <script>
    (() => {{
      const tabButtons = Array.from(document.querySelectorAll('[data-tab-target]'));
      const tabPanels = Array.from(document.querySelectorAll('[data-tab-panel]'));
      tabButtons.forEach((button) => {{
        button.addEventListener('click', () => {{
          const target = button.getAttribute('data-tab-target');
          tabButtons.forEach((b) => b.classList.remove('active'));
          tabPanels.forEach((panel) => {{
            panel.classList.toggle('active', panel.getAttribute('data-tab-panel') === target);
          }});
          button.classList.add('active');
        }});
      }});

      const teamTabButtons = Array.from(document.querySelectorAll('[data-team-tab-target]'));
      const teamPanels = Array.from(document.querySelectorAll('[data-team-panel]'));
      teamTabButtons.forEach((button) => {{
        button.addEventListener('click', () => {{
          const target = button.getAttribute('data-team-tab-target');
          teamTabButtons.forEach((b) => b.classList.remove('active'));
          teamPanels.forEach((panel) => {{
            panel.classList.toggle('active', panel.getAttribute('data-team-panel') === target);
          }});
          button.classList.add('active');
        }});
      }});
    }})();
  </script>
</body>
</html>
"""




